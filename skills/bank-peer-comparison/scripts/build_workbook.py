#!/usr/bin/env python3
"""Render a formatted peer-comparison Excel workbook from the comparison CSVs.

Reads peer_comparison_wide.csv, peer_comparison_stats.csv, and
comparison_metadata.json (from build_comparison.py) and produces a multi-tab
workbook: Summary, Comparison, Stats, Sources. The target bank's row is
highlighted throughout. Values are written as data (already computed from
regulatory source); per-metric peer median/min/max use live Excel formulas so
the sheet recalculates if a user edits a value.

Usage:
    python build_workbook.py --in-dir . --out peer_comparison.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
TARGET_FILL = PatternFill("solid", start_color="FFF2CC")   # soft gold
HEADER_FILL = PatternFill("solid", start_color="1F3864")   # navy
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TARGET_FONT = Font(name=FONT, bold=True, size=10)
BODY_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_KEYS = {  # metrics formatted as percentages
    "roa_pct", "roa_pretax_pct", "roe_pct", "nim_pct", "earning_asset_yield_pct",
    "cost_of_funds_pct", "loan_yield_pct_approx", "deposit_cost_pct_approx",
    "efficiency_pct", "ppnr_to_assets_pct", "nie_to_assets_pct",
    "nir_to_total_revenue_pct", "nir_to_assets_pct", "loans_to_deposits_pct",
    "loans_to_assets_pct", "securities_to_assets_pct", "uninsured_deposits_pct",
    "commercial_loan_pct", "cre_loan_pct", "cre_to_capital_pct", "mortgage_loan_pct",
    "asset_growth_yoy_pct", "loan_growth_yoy_pct", "deposit_growth_yoy_pct",
    "npa_to_assets_pct", "nco_to_loans_pct", "allowance_to_loans_pct", "cet1_pct",
    "tier1_rbc_pct", "total_rbc_pct", "leverage_pct", "equity_to_assets_pct",
    "tce_to_ta_pct", "rotce_pct",
    "target_percentile_vs_peers",
}
DOLLAR_KEYS = {"assets_$000", "net_loans_$000", "deposits_$000", "equity_$000",
               "net_income_ytd_$000", "nir_$000", "deposits_per_branch_$000",
               "assets_per_employee_$000"}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as fh:
        return list(csv.DictReader(fh))


def to_number(v: str) -> Any:
    if v in ("", None):
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_comparison_sheet(ws, wide: list[dict[str, str]], metrics: list[tuple[str, str]]) -> None:
    ws.title = "Comparison"
    ws["A1"] = "Peer Comparison — Bank-Level FDIC Call Report Basis"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(wide) + 2)

    header_row = 3
    ws.cell(row=header_row, column=1, value="Metric")
    # Bank columns; target first if present.
    wide_sorted = sorted(wide, key=lambda r: (r.get("is_target") != "True", r.get("name", "")))
    for j, bank in enumerate(wide_sorted, start=2):
        label = bank["name"]
        if bank.get("is_target") == "True":
            label = f"{label} (target)"
        ws.cell(row=header_row, column=j, value=label)
    median_col = len(wide_sorted) + 2
    ws.cell(row=header_row, column=median_col, value="Peer median")
    style_header(ws, header_row, median_col)

    first_data = header_row + 1
    for i, (key, label) in enumerate(metrics):
        r = first_data + i
        mcell = ws.cell(row=r, column=1, value=label)
        mcell.font = Font(name=FONT, bold=True, size=10)
        mcell.border = BORDER
        peer_cells = []
        for j, bank in enumerate(wide_sorted, start=2):
            val = to_number(bank.get(key, ""))
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = TARGET_FONT if bank.get("is_target") == "True" else BODY_FONT
            cell.border = BORDER
            if bank.get("is_target") == "True":
                cell.fill = TARGET_FILL
            else:
                peer_cells.append(cell.coordinate)
            if key in PCT_KEYS:
                cell.number_format = '0.00"%"'
            elif key in DOLLAR_KEYS:
                cell.number_format = '#,##0;(#,##0);-'
            else:
                cell.number_format = '#,##0'
        # Live peer-median formula across peer (non-target) columns.
        mc = ws.cell(row=r, column=median_col)
        if peer_cells:
            mc.value = f"=MEDIAN({','.join(peer_cells)})"
        mc.font = BODY_FONT
        mc.border = BORDER
        mc.number_format = '0.00"%"' if key in PCT_KEYS else (
            '#,##0;(#,##0);-' if key in DOLLAR_KEYS else '#,##0')

    ws.column_dimensions["A"].width = 34
    for c in range(2, median_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.freeze_panes = ws.cell(row=first_data, column=2)


def build_stats_sheet(ws, stats: list[dict[str, str]]) -> None:
    ws["A1"] = "Target vs Peer Group — Position Summary"
    ws["A1"].font = TITLE_FONT
    cols = ["Metric", "Target", "Peer min", "Peer median", "Peer max", "Target percentile"]
    keys = ["metric_label", "target_value", "peer_min", "peer_median", "peer_max",
            "target_percentile_vs_peers"]
    hr = 3
    for j, c in enumerate(cols, start=1):
        ws.cell(row=hr, column=j, value=c)
    style_header(ws, hr, len(cols))
    for i, row in enumerate(stats):
        r = hr + 1 + i
        for j, (col, key) in enumerate(zip(cols, keys), start=1):
            val = row[key] if key == "metric_label" else to_number(row.get(key, ""))
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if key == "metric_label":
                cell.font = Font(name=FONT, bold=True, size=10)
            elif "pct" in row["metric_key"] and key != "target_percentile_vs_peers":
                cell.number_format = '0.00"%"'
            elif key == "target_percentile_vs_peers":
                cell.number_format = '0"%ile"'
            else:
                cell.number_format = '#,##0'
    ws.column_dimensions["A"].width = 34
    for c in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B4"


def build_summary_sheet(ws, meta: dict[str, Any], wide: list[dict[str, str]]) -> None:
    ws["A1"] = "Bank Peer Comparison"
    ws["A1"].font = TITLE_FONT
    target = meta.get("target", {})
    info = [
        ("Target bank", target.get("name")),
        ("FDIC certificate", target.get("cert")),
        ("State / Fed district", f"{target.get('state')} / {target.get('fed_district_name','')}"),
        ("Specialization", target.get("specialization")),
        ("As-of period", meta.get("as_of")),
        ("Peers in set", len(wide) - 1),
        ("Built (UTC)", meta.get("built_at")),
        ("Basis", "Bank-level FDIC BankFind Financials (Call Report-derived)"),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = BODY_FONT
    row = 3 + len(info) + 1
    ws.cell(row=row, column=1, value="Caveats").font = Font(name=FONT, bold=True, size=11)
    for i, c in enumerate(meta.get("caveats", []), start=row + 1):
        ws.cell(row=i, column=1, value=f"• {c}").font = BODY_FONT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60


def build_sources_sheet(ws, meta: dict[str, Any]) -> None:
    ws["A1"] = "Sources & Methodology"
    ws["A1"].font = TITLE_FONT
    lines = [
        "Primary data: FDIC BankFind Suite APIs (institutions + financials).",
        "  https://api.fdic.gov/banks/institutions",
        "  https://api.fdic.gov/banks/financials",
        "Field dictionary: https://api.fdic.gov/banks/docs",
        "Holding-company context (FR Y-9C): https://www.ffiec.gov/npw and",
        "  https://www.federalreserve.gov/apps/reportingforms/Report/Index/FR_Y-9C",
        "Capital framework reference (PCA categories):",
        "  https://www.fdic.gov/regulations/examinations/enforcement-actions/ch-05.pdf",
        "",
        "Peer selection: asset band + business model (hard gate on SPECGRP/BKCLASS),",
        "Fed district / region as grouping and tie-break, proximity score ranks the set.",
        "Reporting basis: bank-level FDIC CERT data for all members for comparability.",
    ]
    for i, ln in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=ln).font = BODY_FONT
    ws.column_dimensions["A"].width = 90


# Metric order mirrors build_comparison.METRICS.
METRICS = [
    ("assets_$000", "Total assets ($000)"), ("net_loans_$000", "Net loans ($000)"),
    ("deposits_$000", "Deposits ($000)"), ("equity_$000", "Equity ($000)"),
    ("branches", "Branches"), ("employees", "Headcount (FTE)"),
    ("asset_growth_yoy_pct", "Asset growth YoY %"),
    ("loan_growth_yoy_pct", "Loan growth YoY %"),
    ("deposit_growth_yoy_pct", "Deposit growth YoY %"),
    ("net_income_ytd_$000", "Net income YTD ($000)"), ("roa_pct", "ROA %"),
    ("roa_pretax_pct", "Pre-tax ROA %"), ("roe_pct", "ROE %"), ("nim_pct", "NIM %"),
    ("earning_asset_yield_pct", "Earning-asset yield %"),
    ("cost_of_funds_pct", "Cost of funding earning assets %"),
    ("loan_yield_pct_approx", "Loan yield % (approx)"),
    ("deposit_cost_pct_approx", "Deposit cost % (approx)"),
    ("efficiency_pct", "Efficiency ratio %"),
    ("ppnr_to_assets_pct", "PPNR / avg assets %"),
    ("nie_to_assets_pct", "Noninterest expense / avg assets %"),
    ("nir_$000", "Noninterest income YTD ($000)"),
    ("nir_to_total_revenue_pct", "NIR / total revenue %"),
    ("nir_to_assets_pct", "NIR / avg assets %"),
    ("loans_to_deposits_pct", "Loans / deposits %"),
    ("loans_to_assets_pct", "Loans / assets %"),
    ("securities_to_assets_pct", "Securities / assets %"),
    ("uninsured_deposits_pct", "Uninsured deposits %"),
    ("commercial_loan_pct", "C&I loans % of gross"),
    ("cre_loan_pct", "CRE loans % of gross (supervisory)"),
    ("cre_to_capital_pct", "CRE loans / equity capital %"),
    ("mortgage_loan_pct", "1-4 family mortgage % of gross"),
    ("deposits_per_branch_$000", "Deposits per branch ($000)"),
    ("assets_per_employee_$000", "Assets per employee ($000)"),
    ("npa_to_assets_pct", "NPA / assets %"), ("nco_to_loans_pct", "Net charge-offs / loans %"),
    ("allowance_to_loans_pct", "Allowance / gross loans %"), ("cet1_pct", "CET1 %"),
    ("tier1_rbc_pct", "Tier 1 RBC %"), ("total_rbc_pct", "Total RBC %"),
    ("leverage_pct", "Leverage %"), ("equity_to_assets_pct", "Equity / assets %"),
    ("tce_to_ta_pct", "Tangible common equity / tangible assets %"),
    ("rotce_pct", "Return on tangible common equity %"),
]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-dir", default=".")
    ap.add_argument("--out", default="peer_comparison.xlsx")
    args = ap.parse_args()

    in_dir = Path(args.in_dir)
    wide = read_csv(in_dir / "peer_comparison_wide.csv")
    stats = read_csv(in_dir / "peer_comparison_stats.csv")
    meta = json.loads((in_dir / "comparison_metadata.json").read_text())

    wb = Workbook()
    build_summary_sheet(wb.active, meta, wide)
    wb.active.title = "Summary"
    build_comparison_sheet(wb.create_sheet("Comparison"), wide, METRICS)
    build_stats_sheet(wb.create_sheet("Stats"), stats)
    build_sources_sheet(wb.create_sheet("Sources"), meta)
    wb.save(args.out)
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
